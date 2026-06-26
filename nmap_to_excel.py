#!/usr/bin/env python3
"""Convert an nmap XML scan into an Excel workbook.

Sheets:
  Hosts     - one row per host with summary counts (open ports, CVEs, dirs/endpoints)
  Services  - one row per open/filtered port (service/product/version)
  Findings  - one row per notable NSE script result (CVEs, dirs/endpoints, etc.)
"""
import re
import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CVE_RE = re.compile(r'CVE-\d{4}-\d{3,7}', re.IGNORECASE)

# NSE scripts whose results describe discovered directories / endpoints / paths.
DIR_SCRIPTS = {
    'http-enum', 'http-robots.txt', 'http-config-backup',
    'http-default-accounts', 'http-vhosts', 'http-backup-finder',
}


def classify(script_id, output):
    """Return a finding category for a script result, or None to skip it."""
    sid = (script_id or '').lower()
    has_cve = bool(CVE_RE.search(output or '')) or 'cve' in sid
    is_vuln = 'vuln' in sid or has_cve
    if is_vuln:
        return 'CVE / Vulnerability'
    if sid in DIR_SCRIPTS:
        return 'Directory / Endpoint'
    return None


def first_addr(host, kinds):
    for a in host.findall('address'):
        if a.get('addrtype') in kinds:
            return a.get('addr')
    return ''


def parse(xml_path):
    hosts_rows, service_rows, finding_rows = [], [], []

    for _, host in ET.iterparse(xml_path, events=('end',)):
        if host.tag != 'host':
            continue

        status = host.find('status')
        if status is not None and status.get('state') == 'down':
            host.clear()
            continue

        ip = first_addr(host, ('ipv4', 'ipv6'))
        mac = first_addr(host, ('mac',))
        names = [h.get('name') for h in host.findall('hostnames/hostname') if h.get('name')]
        hostname = names[0] if names else ''

        os_match = host.find('os/osmatch')
        os_name = os_match.get('name') if os_match is not None else ''

        open_ports = 0
        host_cves = set()
        dir_count = 0

        # host-level scripts (hostscript)
        script_iters = list(host.findall('hostscript/script'))

        for port in host.findall('ports/port'):
            state_el = port.find('state')
            state = state_el.get('state') if state_el is not None else ''
            proto = port.get('protocol', '')
            portid = port.get('portid', '')

            svc = port.find('service')
            sname = svc.get('name', '') if svc is not None else ''
            product = svc.get('product', '') if svc is not None else ''
            version = svc.get('version', '') if svc is not None else ''
            extra = svc.get('extrainfo', '') if svc is not None else ''

            if state == 'open':
                open_ports += 1

            service_rows.append([
                ip, hostname, portid, proto, state,
                sname, product, version, extra,
            ])

            for script in port.findall('script') + script_iters:
                sid = script.get('id', '')
                out = script.get('output', '') or ''
                cat = classify(sid, out)
                if cat is None:
                    continue
                cves = sorted(set(c.upper() for c in CVE_RE.findall(out)))
                if cat == 'CVE / Vulnerability':
                    host_cves.update(cves)
                else:
                    dir_count += 1
                finding_rows.append([
                    ip, hostname,
                    f"{portid}/{proto}" if portid else '',
                    cat, sid, ', '.join(cves),
                    out.strip(),
                ])
            script_iters = []  # only attach hostscripts once

        hosts_rows.append([
            ip, hostname, mac, os_name, open_ports,
            len(host_cves), dir_count,
            ', '.join(sorted(host_cves)),
        ])
        host.clear()

    return hosts_rows, service_rows, finding_rows


HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF')
CVE_FILL = PatternFill('solid', fgColor='F8CBAD')
DIR_FILL = PatternFill('solid', fgColor='C6E0B4')


def write_sheet(ws, headers, rows, wrap_cols=(), max_width=70):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center')
    for r in rows:
        ws.append(r)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    # column widths
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        longest = len(str(headers[c - 1]))
        for r in rows:
            v = r[c - 1] if c - 1 < len(r) else ''
            longest = max(longest, len(str(v).split('\n')[0]))
        ws.column_dimensions[letter].width = min(longest + 2, max_width)
    for c in wrap_cols:
        letter = get_column_letter(c)
        for row in range(2, len(rows) + 2):
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical='top')


def main(xml_path, out_path):
    hosts_rows, service_rows, finding_rows = parse(xml_path)

    wb = Workbook()

    ws = wb.active
    ws.title = 'Hosts'
    write_sheet(ws, [
        'IP', 'Hostname', 'MAC', 'OS', 'Open Ports',
        'CVE Count', 'Dir/Endpoint Count', 'CVEs',
    ], hosts_rows, wrap_cols=(8,))

    ws2 = wb.create_sheet('Services')
    write_sheet(ws2, [
        'IP', 'Hostname', 'Port', 'Protocol', 'State',
        'Service', 'Product', 'Version', 'Extra Info',
    ], service_rows)

    ws3 = wb.create_sheet('Findings')
    write_sheet(ws3, [
        'IP', 'Hostname', 'Port', 'Category', 'Script', 'CVEs', 'Details',
    ], finding_rows, wrap_cols=(7,))
    # color-code findings by category
    for row in range(2, len(finding_rows) + 2):
        cat = ws3.cell(row=row, column=4).value
        fill = CVE_FILL if cat == 'CVE / Vulnerability' else DIR_FILL
        ws3.cell(row=row, column=4).fill = fill

    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  Hosts:    {len(hosts_rows)}")
    print(f"  Services: {len(service_rows)}")
    print(f"  Findings: {len(finding_rows)} "
          f"({sum(1 for r in finding_rows if r[3]=='CVE / Vulnerability')} CVE/vuln, "
          f"{sum(1 for r in finding_rows if r[3]=='Directory / Endpoint')} dir/endpoint)")


if __name__ == '__main__':
    xml = sys.argv[1] if len(sys.argv) > 1 else 'OUTPUT_FILE_OUT.xml'
    out = sys.argv[2] if len(sys.argv) > 2 else 'OUTPUT_FILE_OUT.xlsx'
    main(xml, out)
